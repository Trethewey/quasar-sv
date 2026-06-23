"""Sample-level metadata loader for cohort reports.

Reads the user's `cohort_metadata.xlsx` (or any compatible TSV) and maps
sample identifiers used in fusion-call TSVs to clinical labels (cell line
name, cohort subtype, aSHM expectation, est. coverage).
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError:
    openpyxl = None   # XLSX support optional; TSV path always works


@dataclass
class SampleMetadata:
    sample_id: str          # the key as used in fusion-call TSVs (e.g. 'ERR9188549_Karpas1106P')
    run: str = ""           # ENA/SRA/DDBJ run accession
    cell_line: str = ""     # e.g. 'Karpas1106P'
    cohort: str = ""        # 'PMBL', 'ATLL', 'DLBCL', ...
    ashm_expected: str = "" # 'pos' / 'neg' / ''
    coverage: float = 0.0
    project: str = ""       # BioProject accession


def load_cohort_metadata_xlsx(
    path: str,
    sample_id_format: str = "{run}_{title}",
) -> list[SampleMetadata]:
    """Parse the cohort_metadata.xlsx file (or compatible XLSX with matching headers)."""
    if openpyxl is None:
        raise ImportError("openpyxl required for XLSX loading")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    out: list[SampleMetadata] = []
    for r in ws.iter_rows(values_only=True):
        if header is None:
            header = list(r)
            continue
        rec = dict(zip(header, r))
        run = (rec.get("Run") or "").strip()
        if not run:
            continue
        title = (rec.get("Sample title") or "").replace(" ", "_")
        sid = sample_id_format.format(run=run, title=title) if title else run
        out.append(SampleMetadata(
            sample_id=sid,
            run=run,
            cell_line=(rec.get("Sample title") or "").strip(),
            cohort=(rec.get("Cohort") or "").strip(),
            ashm_expected=(rec.get("aSHM expected") or "").strip(),
            coverage=_as_float(rec.get("Est. coverage (×, 3 Gb genome)")),
            project=(rec.get("Project") or "").strip(),
        ))
    return out


def _as_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def build_metadata_index(items: Iterable[SampleMetadata]) -> dict[str, SampleMetadata]:
    return {m.sample_id: m for m in items}


# Cohort label → lineage prior. B-cell lymphomas dominate; T-cell labelled
# explicitly. Anything else falls through to "any" (no lineage filter applied).
_TCELL_COHORTS = {
    "ATLL", "PTCL", "ALCL", "AITL", "CTCL", "T-CELL NHL", "T-CELL_NHL",
    "T-LBL", "TLBL", "T-ALL", "TALL", "TCL",
}
_BCELL_COHORTS = {
    "DLBCL", "PMBL", "FL", "MCL", "MALT", "MZL", "BL", "LPL", "CLL",
    "DH-DLBCL", "DH_DLBCL", "TH-DLBCL", "PCNSL", "EMZL", "NMZL", "SMZL",
}


def cohort_to_lineage(cohort: str) -> str:
    """Map a cohort label to ``"B"`` / ``"T"`` / ``"any"``."""
    c = (cohort or "").upper().strip()
    if c in _TCELL_COHORTS:
        return "T"
    if c in _BCELL_COHORTS:
        return "B"
    return "any"


def lineage_index_from_metadata(items: Iterable[SampleMetadata]) -> dict[str, str]:
    """Return ``{sample_id: "B" | "T" | "any"}`` derived from each row's cohort."""
    return {m.sample_id: cohort_to_lineage(m.cohort) for m in items}
