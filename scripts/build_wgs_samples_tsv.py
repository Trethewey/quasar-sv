"""Build samples.tsv for wgs_cohort_batch.sh from W:\\WGS_data\\cohort_metadata.xlsx.

Output format (tab-separated, no header):
    <sample_id>\t<cram_path>\t<reference_fa>

Reference path is supplied by --ref because the CRAM may have been encoded with
a specific GRCh38 build that the user must point at.
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--metadata",
                    required=True,
                    help="Cohort metadata XLSX with one row per sample")
    ap.add_argument("--cram-root", required=True,
                    help="Root directory containing per-BioProject CRAM sub-trees")
    ap.add_argument("--ref", required=True,
                    help="Path to the GRCh38 reference matched to the CRAMs")
    ap.add_argument("--output", required=True)
    ap.add_argument("--cohort-filter",
                    help="Only include rows where the Cohort column matches this (e.g. 'PMBL')")
    ap.add_argument("--wsl-prefix", default="",
                    help="Optional drive-letter -> WSL mount remap (e.g. 'W:' -> '/mnt/w')")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.metadata, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    rows: list[dict[str, str]] = []
    for r in ws.iter_rows(values_only=True):
        if header is None:
            header = list(r)
            continue
        rec = dict(zip(header, r))
        if args.cohort_filter and rec.get("Cohort") != args.cohort_filter:
            continue
        rows.append(rec)

    out_lines = []
    for rec in rows:
        prj = rec.get("Project") or ""
        run = rec.get("Run") or ""
        title = rec.get("Sample title") or ""
        if not prj or not run:
            continue
        cram_local = Path(args.cram_root) / prj / "cram" / f"{run}.cram"
        cram_wsl = str(cram_local).replace("W:", args.wsl_prefix).replace("\\", "/")
        sample_id = f"{run}_{title.replace(' ', '_')}" if title else run
        out_lines.append(f"{sample_id}\t{cram_wsl}\t{args.ref}")

    Path(args.output).write_text("\n".join(out_lines) + "\n", encoding="utf-8")
    print(f"wrote {len(out_lines)} samples -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
